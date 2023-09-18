import json
import openpyxl
from datetime import datetime as dt
from gui.create_gui import make_msgbox
from tools.calculate_date import calc_date
from tools.create_sheet_cell import create_cell
from register_teacher.get_teacher_row import get_row
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from files_paths.files import management_student_file, teacher_file
from monthly_sheet.create_teacher_monthly_columns import create_teacher_monthly_columns


def add_new_sheet():
    file_name = management_student_file
    wb = openpyxl.load_workbook(file_name)
    db_sheet = wb["PythonDatabase"]

    change_month = db_sheet["B1"].value
    if change_month == "change month":
        try:
            teacher_file_name = teacher_file
            teacher_wb = openpyxl.load_workbook(teacher_file_name)
            teacher_db_sheet = teacher_wb["PythonDatabase"]

            orange_bg = PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            border_style = Side(border_style="thin", color="000000")
            black = Border(
                left=border_style,
                right=border_style,
                top=border_style,
                bottom=border_style,
            )

            months = {
                1: "JANUARY",
                2: "FEBRUARY",
                3: "MARCH",
                4: "APRIL",
                5: "MAY",
                6: "JUNE",
                7: "JULY",
                8: "AUGUST",
                9: "SEPTEMBER",
                10: "OCTOBER",
                11: "NOVEMBER",
                12: "DECEMBER",
            }

            create_teacher_monthly_columns(
                teacher_db_sheet, teacher_wb, months, orange_bg, bold, center, black
            )

            db_maxrow = db_sheet.max_row + 1
            for row in range(2, db_maxrow):
                student_data = db_sheet[f"A{row}"].value.split("+")
                classes = json.loads(student_data[1].replace("'", '"'))
                student_sheet = wb[student_data[0]]
                student_month_row = student_sheet.max_row + 2

                student_sheet.merge_cells(
                    start_row=student_month_row,
                    end_row=student_month_row,
                    start_column=2,
                    end_column=7,
                )
                student_sheet_cell = student_sheet.cell(
                    row=student_month_row,
                    column=2,
                    value=f"{months[dt.now().month]} CLASSES",
                )
                student_sheet_cell.fill = orange_bg
                student_sheet_cell.font = bold
                student_sheet_cell.alignment = center
                for col in range(2, 8):
                    student_sheet.cell(row=student_month_row, column=col).border = black

                column_data = [
                    f"Pay Day 10/{dt.now().month}",
                    "Date",
                    "Teacher",
                    "Subject",
                    "Hour",
                    "Attendance",
                ]
                student_columns_row = student_month_row + 1
                for num, col_value in enumerate(column_data, start=2):
                    cell = create_cell(
                        student_sheet,
                        student_columns_row,
                        num,
                        col_value,
                        black,
                        center,
                    )
                    cell.font = bold

                days = calc_date(
                    student_sheet["E4"].value.replace(" e ", ", ").split(", ")
                )
                student_row = student_columns_row + 1

                create_cell(student_sheet, student_row, 2, "Not Paid", black, center)

                teacher_key = {}
                for num in classes:
                    key = f"{classes[num]['subject']}-{classes[num]['day']}-{classes[num]['class_hours']}-{classes[num]['teacher']}-{classes[num]['price']}"
                    teacher_key[key] = {}
                    teacher_key[key]["subject"] = classes[num]["subject"]
                    teacher_key[key]["teacher"] = classes[num]["teacher"]
                    teacher_key[key]["class_hours"] = classes[num]["class_hours"]
                    teacher_key[key]["price"] = classes[num]["price"]
                    teacher_key[key]["day"] = classes[num]["day"]
                    teacher_key[key]["name"] = classes[num]["name"]
                    teacher_key[key]["chosen_plan"] = classes[num]["chosen_plan"]

                adder = 0
                for row_num, day in enumerate(days, start=student_row):
                    for key in teacher_key:
                        if day.split(", ")[1] in key:
                            row = row_num + adder
                            adder += 1
                            create_cell(student_sheet, row, 3, day, black, center)
                            create_cell(
                                student_sheet,
                                row,
                                4,
                                teacher_key[key]["teacher"],
                                black,
                                center,
                            )
                            create_cell(
                                student_sheet,
                                row,
                                5,
                                teacher_key[key]["subject"],
                                black,
                                center,
                            )
                            create_cell(
                                student_sheet,
                                row,
                                6,
                                teacher_key[key]["class_hours"],
                                black,
                                center,
                            )
                            create_cell(student_sheet, row, 7, "", black, center)

                            try:
                                teacher_sheet = teacher_wb[
                                    f"{teacher_key[key]['teacher']}"
                                ]
                            except Exception:
                                raise ValueError(
                                    f"{teacher_key[key]['teacher']}'s sheet doesn't exist!"
                                )
                            teacher_row = get_row(teacher_sheet)
                            if teacher_row == False:
                                raise ValueError(
                                    f"Error in {teacher_key[key]['teacher']}'s sheet! Try again."
                                )
                            create_cell(
                                teacher_sheet, teacher_row, 2, day, black, center
                            )
                            create_cell(
                                teacher_sheet,
                                teacher_row,
                                3,
                                teacher_key[key]["class_hours"],
                                black,
                                center,
                            )
                            create_cell(
                                teacher_sheet,
                                teacher_row,
                                4,
                                teacher_key[key]["name"],
                                black,
                                center,
                            )
                            create_cell(
                                teacher_sheet,
                                teacher_row,
                                5,
                                teacher_key[key]["subject"],
                                black,
                                center,
                            )
                            create_cell(
                                teacher_sheet,
                                teacher_row,
                                6,
                                teacher_key[key]["chosen_plan"],
                                black,
                                center,
                            )
                            create_cell(
                                teacher_sheet, teacher_row, 7, "", black, center
                            )
                            create_cell(
                                teacher_sheet,
                                teacher_row,
                                8,
                                f"$ {float(teacher_key[key]['price']):.2f}",
                                black,
                                center,
                            )
                            create_cell(
                                teacher_sheet, teacher_row, 9, "", black, center
                            )
            db_sheet["B1"].value = "do not change month"
            wb.save(file_name)
            teacher_wb.save(teacher_file_name)
            wb.close()
            teacher_wb.close()
            make_msgbox("SUCCESS", "The sheet was created successfully!", "check")
            return True
        except Exception as e:
            make_msgbox(
                "ERROR",
                f"An error occured when trying to create the sheets!\n\nERROR:{e}\n\nTry again.",
                "cancel",
            )
            return False
