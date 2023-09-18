import openpyxl
from datetime import datetime as dt
from gui.create_gui import make_msgbox
from tools.calculate_date import calc_date
from tools.create_sheet_cell import create_cell
from register_teacher.get_teacher_row import get_row
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from files_paths.files import management_student_file, teacher_file


def create_sheet(
    name, classes, classes_to_str, chosen_plan, initial_date, expiration_date
):
    warning_msg = make_msgbox(
        "ERROR",
        "For any information to be added to an Excel file, it must be closed to prevent possible failures and/or errors. "
        "So make sure Excel is closed!\n\nDo you want to continue?",
        "question",
        "Continue",
        "Cancel",
    )

    if warning_msg.get() == "Continue":
        try:
            teacher_file_name = teacher_file
            teacher_wb = openpyxl.load_workbook(teacher_file_name)

            file_name = management_student_file
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.create_sheet(name)

            database_sheet = wb["PythonDatabase"]
            db_row = database_sheet.max_row

            orange_bg = PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            border_style = Side(border_style="thin", color="000000")
            black = Border(
                left=border_style,
                right=border_style,
                top=border_style,
                bottom=border_style,
            )

            months = {
                1: "JANUARY",
                2: "FEBRUARY",
                3: "MARCH",
                4: "APRIL",
                5: "MAY",
                6: "JUNE",
                7: "JULY",
                8: "AUGUST",
                9: "SEPTEMBER",
                10: "OCTOBER",
                11: "NOVEMBER",
                12: "DECEMBER",
            }

            sheet.merge_cells(start_row=2, end_row=2, start_column=2, end_column=5)
            sheet_cell = sheet.cell(row=2, column=2, value=f"{name}".upper())
            sheet_cell.fill = orange_bg
            sheet_cell.font = bold
            sheet_cell.alignment = center
            for col in range(2, 6):
                sheet.cell(row=2, column=col).border = black

            sheet.merge_cells(start_row=6, end_row=6, start_column=2, end_column=7)
            sheet_cell = sheet.cell(
                row=6, column=2, value=f"{months[dt.now().month]} CLASSES"
            )
            sheet_cell.fill = orange_bg
            sheet_cell.font = bold
            sheet_cell.alignment = center
            for col in range(2, 8):
                sheet.cell(row=6, column=col).border = black

            main_data = {
                "column": [
                    "Cicle",
                    "Plan",
                    "N° Classes/Week",
                    "Classes Days",
                ],
                "row": [
                    f"{initial_date} to {expiration_date}",
                    f"{chosen_plan}",
                    f"{str(classes_to_str[3])}",
                    f"{classes_to_str[4]}",
                ],
            }

            for num, col_value in enumerate(main_data["column"], start=2):
                cell = create_cell(sheet, 3, num, col_value, black, center)
                cell.font = bold

            for num, row_value in enumerate(main_data["row"], start=2):
                create_cell(sheet, 4, num, row_value, black, center)

            column_data = [
                f"Pay Day 10/{dt.now().month}",
                "Date",
                "Teacher",
                "Subject",
                "Hour",
                "Attendance",
            ]

            days = calc_date(classes_to_str[4].replace(" e ", ", ").split(", "))

            teacher_key = {}
            for num in classes:
                key = f'{classes[num]["subject"]}-{classes[num]["day"]}-{classes[num]["class_hours"]}-{classes[num]["teacher"]}-{classes[num]["price"]}'
                teacher_key[key] = {}
                teacher_key[key]["subject"] = classes[num]["subject"]
                teacher_key[key]["teacher"] = classes[num]["teacher"]
                teacher_key[key]["class_hours"] = classes[num]["class_hours"]
                teacher_key[key]["price"] = classes[num]["price"]
                teacher_key[key]["day"] = classes[num]["day"]
                teacher_key[key]["name"] = name
                teacher_key[key]["chosen_plan"] = chosen_plan
            database_sheet[f"A{db_row + 1}"] = f"{name}+{teacher_key}"

            for num, col_value in enumerate(column_data, start=2):
                cell = create_cell(sheet, 7, num, col_value, black, center)
                cell.font = bold

            create_cell(sheet, 8, 2, "Not paid", black, center)

            adder = 0
            for row_num, day in enumerate(days, start=8):
                for key in teacher_key:
                    if day.split(", ")[1] in key:
                        row = row_num + adder
                        adder += 1
                        create_cell(sheet, row, 3, day, black, center)
                        create_cell(
                            sheet, row, 4, teacher_key[key]["teacher"], black, center
                        )
                        create_cell(
                            sheet, row, 5, teacher_key[key]["subject"], black, center
                        )
                        create_cell(
                            sheet,
                            row,
                            6,
                            teacher_key[key]["class_hours"],
                            black,
                            center,
                        )
                        create_cell(sheet, row, 7, "", black, center)

                        try:
                            teacher_sheet = teacher_wb[f"{teacher_key[key]['teacher']}"]
                        except Exception:
                            raise ValueError(
                                f"The {teacher_key[key]['teacher']}'s sheet doesn't exist!"
                            )
                        teacher_row = get_row(teacher_sheet)
                        if teacher_row == False:
                            raise ValueError(
                                f"An error occured when trying to read {teacher_key[key]['teacher']}'s sheet! Try again."
                            )
                        create_cell(teacher_sheet, teacher_row, 2, day, black, center)
                        create_cell(
                            teacher_sheet,
                            teacher_row,
                            3,
                            teacher_key[key]["class_hours"],
                            black,
                            center,
                        )
                        create_cell(teacher_sheet, teacher_row, 4, name, black, center)
                        create_cell(
                            teacher_sheet,
                            teacher_row,
                            5,
                            teacher_key[key]["subject"],
                            black,
                            center,
                        )
                        create_cell(
                            teacher_sheet, teacher_row, 6, chosen_plan, black, center
                        )
                        create_cell(teacher_sheet, teacher_row, 7, "", black, center)
                        create_cell(
                            teacher_sheet,
                            teacher_row,
                            8,
                            f"$ {float(teacher_key[key]['price']):.2f}",
                            black,
                            center,
                        )
                        create_cell(teacher_sheet, teacher_row, 9, "", black, center)
            try:
                wb.save(file_name)
                wb.close()
                teacher_wb.save(teacher_file_name)
                teacher_wb.close()
                make_msgbox(
                    "SUCCESS",
                    f"The student {name}'s sheet was created succesfully!\n\nThe teacher's data was inserted to teacher's sheets!",
                    "check",
                )
            except Exception as e:
                make_msgbox(
                    "ERROR",
                    f"The student {name}'s sheet was not created because an error occured.\nClose all the Excel's tabs!\nTry again.",
                    "cancel",
                )
        except Exception as e:
            make_msgbox(
                "ERROR",
                f"The student {name}'s sheet has not been created!\n\nERROR: {str(e)}\n\nTry again.",
                "cancel",
            )
    else:
        make_msgbox("WARNING", "Execution canceled!", "info")
