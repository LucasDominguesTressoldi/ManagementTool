import openpyxl
from datetime import datetime as dt
from gui.create_gui import make_msgbox
from files_paths.files import teacher_file
from tools.create_sheet_cell import create_cell
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill


def create_teacher(name):
    warning_msg = make_msgbox(
        "ERROR",
        "For any information to be added to an Excel file, it must be closed to prevent possible failures and/or errors. "
        "So make sure Excel is closed!\n\nDo you want to continue?",
        "question",
        "Continue",
        "Cancel",
    )

    if warning_msg.get() == "Continue":
        try:
            file_name = teacher_file

            wb = openpyxl.load_workbook(file_name)
            sheet = wb.create_sheet(f"{name}")

            database_sheet = wb["PythonDatabase"]
            db_row = database_sheet.max_row
            database_sheet[f"A{db_row + 1}"] = name

            orange_bg = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid"
            )
            bold = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            border_style = Side(border_style="thin", color="000000")
            black = Border(
                left=border_style,
                right=border_style,
                top=border_style,
                bottom=border_style,
            )

            months = {
                1: "JANUARY",
                2: "FEBRUARY",
                3: "MARCH",
                4: "APRIL",
                5: "MAY",
                6: "JUNE",
                7: "JULY",
                8: "AUGUST",
                9: "SEPTEMBER",
                10: "OCTOBER",
                11: "NOVEMBER",
                12: "DECEMBER",
            }

            sheet.merge_cells(start_row=2, end_row=2, start_column=2, end_column=9)
            sheet_cell = sheet.cell(
                row=2, column=2, value=f"{months[dt.now().month]}".upper()
            )
            sheet_cell.fill = orange_bg
            sheet_cell.font = bold
            sheet_cell.alignment = center
            for col in range(2, 10):
                sheet.cell(row=2, column=col).border = black

            main_data = {
                "column": [
                    "Date",
                    "Hour",
                    "Student",
                    "Subject",
                    "Plan",
                    "Attendance",
                    "Price h/c",
                    "Observation",
                ]
            }

            for num, col_value in enumerate(main_data["column"], start=2):
                cell = create_cell(sheet, 3, num, col_value, black, center)
                cell.font = bold

            try:
                wb.save(file_name)
                wb.close()
                make_msgbox(
                    "SUCCESS",
                    f"The teacher {name} was created succesfully!",
                    "check",
                )
            except Exception as e:
                make_msgbox(
                    "ERROR",
                    f"The teacher {name} was not created because an error occured when trying to save it.\nClose all the Excel's tabs!\nTry again.",
                    "cancel",
                )
        except Exception as e:
            make_msgbox(
                "ERROR",
                f"The teacher {name} has NOT been created!\n\nERROR: {str(e)}\n\nTry again.",
                "cancel",
            )
    else:
        make_msgbox("WARNING", "Execution canceled!", "info")
