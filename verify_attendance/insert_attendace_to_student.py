import openpyxl
from gui.create_gui import make_msgbox
from datetime import date as dte, datetime as dt
from files_paths.files import management_student_file, teacher_file
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill


def update_attendance(student_name, chosen_attendance_opt, day, month, subject, hour):
    try:
        teacher_file_name = teacher_file
        teacher_wb = openpyxl.load_workbook(teacher_file_name)

        file_name = management_student_file
        wb = openpyxl.load_workbook(file_name)

        try:
            sheet = wb[student_name]
        except:
            make_msgbox("WARNING", "Student not found!\nTry again.", "warning")
            return False

        date = dte(dt.now().year, int(month), int(day)).strftime("%d/%m/%Y")

        blue_bg = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        yellow_bg = PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        )
        red_bg = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )
        white_font = Font(color="FFFFFF")
        black_font = Font(color="000000")
        center = Alignment(horizontal="center", vertical="center")
        border_style = Side(border_style="thin", color="000000")
        black_border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style,
        )

        cell_teacher = ""
        file_changed = False
        for row in sheet.iter_rows(min_row=8, min_col=3, max_col=3):
            cell_value = row[0].value
            if cell_value and isinstance(cell_value, str):
                cell_date = cell_value.split(", ")[0]
                cell_teacher = sheet[f"D{row[0].row}"].value
                cell_subject = sheet[f"E{row[0].row}"].value
                cell_hour = sheet[f"F{row[0].row}"].value

                if (
                    cell_date.strip() == date.strip()
                    and cell_subject.strip() == subject.strip()
                    and cell_hour.strip() == hour.strip()
                ):
                    attendance_cell = sheet.cell(
                        row=row[0].row, column=7, value=chosen_attendance_opt
                    )
                    attendance_cell.font = black_font
                    attendance_cell.alignment = center
                    attendance_cell.border = black_border
                    if chosen_attendance_opt == "Done":
                        attendance_cell.fill = blue_bg
                        attendance_cell.font = white_font
                    elif chosen_attendance_opt == "Pending":
                        attendance_cell.fill = yellow_bg
                    elif chosen_attendance_opt == "Missed":
                        attendance_cell.fill = red_bg

                    if cell_teacher != "":
                        teacher_sheet = teacher_wb[f"{cell_teacher}"]
                    else:
                        make_msgbox(
                            "WARNING",
                            "Teacher not found!\nTry again.",
                            "warning",
                        )
                        return False

                    for row in teacher_sheet.iter_rows(min_row=4, min_col=2, max_col=2):
                        cell_value = row[0].value
                        if cell_value and isinstance(cell_value, str):
                            cell_date = cell_value.split(", ")[0]
                            cell_hour = teacher_sheet[f"C{row[0].row}"].value
                            cell_student = teacher_sheet[f"D{row[0].row}"].value
                            cell_subject = teacher_sheet[f"E{row[0].row}"].value

                            if (
                                cell_date.strip() == date.strip()
                                and cell_hour.strip() == hour.strip()
                                and cell_student.strip().lower()
                                == student_name.strip().lower()
                                and cell_subject.strip().lower()
                                == subject.strip().lower()
                            ):
                                file_changed = True
                                attendance_cell = teacher_sheet.cell(
                                    row=row[0].row,
                                    column=7,
                                    value=chosen_attendance_opt,
                                )
                                attendance_cell.font = black_font
                                attendance_cell.alignment = center
                                attendance_cell.border = black_border
                                if chosen_attendance_opt == "Done":
                                    attendance_cell.fill = blue_bg
                                    attendance_cell.font = white_font
                                elif chosen_attendance_opt == "Pending":
                                    attendance_cell.fill = yellow_bg
                                elif chosen_attendance_opt == "Missed":
                                    attendance_cell.fill = red_bg

        try:
            if file_changed == True:
                wb.save(file_name)
                wb.close()
                teacher_wb.save(teacher_file_name)
                teacher_wb.close()
                make_msgbox(
                    "SUCCESS",
                    f"The student {student_name} attendance was modified to {chosen_attendance_opt}!\n\nThe teachers' attendance column was automatically filled in.",
                    "check",
                )
            else:
                make_msgbox(
                    "ERROR",
                    f"The data provided was not found!",
                    "cancel",
                )
        except Exception as e:
            make_msgbox(
                "ERROR",
                "The attendance has NOT been changed because the file is open!\nPlease, close it and try again.",
                "cancel",
            )
    except Exception as e:
        make_msgbox(
            "ERROR",
            f"ERROR: The attendance was not updated!\n\nERROR: {e}\n\nTry again.",
            "cancel",
        )
